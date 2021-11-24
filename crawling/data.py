# JSON
import urllib.request as req
import json
from bs4 import BeautifulSoup as bs
url = "https://api.github.com/repositories"
# resp = req.urlopen( url ).read().decode( "utf-8" )
req.urlretrieve( url, "github.json" )
print( "저장했습니다." );

items = json.load( open( "github.json", "rt", encoding="utf-8" ) )
for item in items :
    print( item["name"] + " : " + item["owner"]["login"] )      # dictionary
    
# CSV Comma Separate Values
import csv, codecs
iris = codecs.open( "iris.csv", "r", encoding="utf-8" ).read()
# print( iris )

rows = iris.split( "\n" )
data = []
print( rows )
for row in rows :
    columns = row.split(",")
    data.append( columns )
for i in range(0, len(data)) :
    for j in range(0, len(data[i])):
        print(data[i][j] + "\t")
    print()
    
file = codecs.open("iris1.csv","w",encoding = "utf-8")
writer = csv.writer(file, delimiter=",")
for d in data :
    writer.writerow(d)
file.close()

# Excel
# pip install openpyxl 
import openpyxl as px
wb = px.load_workbook( "stat_100701.xlsx" )
# print( wb.get_sheet_names() )
# print( wb.get_sheet_by_name( "Sheet0" ) )       # 워크시트 선택
sheet = wb.active

data = []
for row in sheet.rows :
     data.append( [ row[0].value, row[9].value] )

for i in range( len(data) ) :
    print( data[i], end="\t" );     
print()    

w = px.Workbook()
wsheet = w.active
wsheet.title = "지역별인구"
for i, da in enumerate(data) :
    wsheet.cell(i+1, column=1, value=da[0])
    wsheet.cell(i+1, column=2, value=da[1])
w.save("population.xlsx")
print("저장했습니다.")

